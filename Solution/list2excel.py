# -*- coding: utf-8 -*-
"""
Created on Tue May 17 18:44:32 2016

@author: akshatshah
"""

from openpyxl import Workbook

def py2excel(excel_format):
    wb = Workbook()    
    # grab the active worksheet
    ws = wb.active
    ws['B1'] = 0
    ws['C1'] = 1
    ws['D1'] = 2
    ws['E1'] = 3
    ws['F1'] = 4
    ws['G1'] = 5
    ws['H1'] = 6
    ws['I1'] = 7
    for item in excel_format:
        ws.append(item)
    wb.save("Formatted_Parsed_Output.xlsx")